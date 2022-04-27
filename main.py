import requests
from bs4 import BeautifulSoup
import os
import openpyxl


def abrir_excel():
    # endereco_arquivo = r"D:\projetos\pessoal\sistema_paim" \
                       # r"\Documentação Atualizador Publicações\publicacoes_torre_sdco.xlsx"
    # endereco_arquivo = r"Z:\publicacoes_torre_sdco.xlsx"
    endereco_arquivo = r'D:\Jocimar\TWR_SDCO_PROJETOS\publicacoes_torre_sdco.xlsx'
    return openpyxl.open(endereco_arquivo)


def retorna_indices(arquivo_):
    coluna = 2
    coluna_tipo = arquivo_.worksheets[1].cell(row=1, column=coluna).value
    coluna_numero = arquivo_.worksheets[1].cell(row=2, column=coluna).value
    coluna_data = arquivo_.worksheets[1].cell(row=3, column=coluna).value
    coluna_atualizada = arquivo_.worksheets[1].cell(row=4, column=coluna).value
    return [coluna_tipo, coluna_numero, coluna_data, coluna_atualizada]


def total_lista_documento(arquivo_, coluna):
    total_linha = 0
    for i in range(6, 100000):
        valor_celula = arquivo_.worksheets[0].cell(row=i, column=coluna).value
        if valor_celula is None:
            break
        total_linha += 1
    return total_linha


def progress_bar(progress, total):
    porcent = 100 * (progress / float(total))
    bar = "█" * int(porcent) + "-" * (100 - int(porcent))
    print(f"\r|{bar}| {porcent:.2f}%", end='\r')


def verificador_documentacao_atualizado(arquivo_):
    c_tipo, c_numero, c_data, c_atualizada = retorna_indices(arquivo_)
    docs_desatualizado = list()
    total_linha = total_lista_documento(arquivo_, c_tipo)
    progress_bar(0, total_linha)
    for i in range(6, total_linha+6):
        documento_tipo = arquivo_.worksheets[0].cell(row=i, column=c_tipo).value
        if documento_tipo is None:
            break
        documento_numero = arquivo_.worksheets[0].cell(row=i, column=c_numero).value
        if "aic" in documento_tipo or "AIC" in documento_tipo:
            documento_numero = documento_numero.replace("/", "")
        documento_data = arquivo_.worksheets[0].cell(row=i, column=c_data).value
        documento_atualizado = arquivo_.worksheets[0].cell(row=i, column=c_atualizada).value
        data_formatada = converte_data(documento_data)

        # print(f"Documento: {documento_tipo} {documento_numero} é da data de {data_formatada} "
        #       f"e está atualizada {documento_atualizado}")
        doc = verificador_documentacao_ais(arquivo_, [documento_tipo, documento_numero, data_formatada],
                                     [i, c_atualizada])
        if doc != "":
            docs_desatualizado.append(doc)
        progress_bar(i-5, total_linha)
    return docs_desatualizado


def converte_data(data):
    data_separada = [data.day, data.month, data.year]
    if int(data_separada[0]) < 10:
        data_separada[0] = f"0{data_separada[0]}"
    if int(data_separada[1]) < 10:
        data_separada[1] = f"0{data_separada[1]}"
    return f"{data_separada[0]}/{data_separada[1]}/{data_separada[2]}"


def trata_data_site(data):
    data_lista = data.split(" de ")
    data_formatada = f"{data_lista[0]}/{tranforma_nome_mes_numero(data_lista[1])}/{data_lista[2]}"
    return data_formatada


def tranforma_nome_mes_numero(mes_):
    lista_meses = ["janeiro", "fevereiro", "março", "abril",
                   "maio", "junho", "julho", "agosto",
                   "setembro", "outubro", "novembro", "dezembro"]
    mes = lista_meses.index(mes_.lower())+1
    if mes < 10:
        mes = f'0{mes}'
    return mes


def verificador_documentacao_ais(arq, regulamento_, lista_indices):
    documento_desatualizado = ""
    # print(regulamento_)
    regulamento = f"{regulamento_[0].lower()}-{regulamento_[1]}"
    # print(regulamento)
    r = requests.get(f"https://publicacoes.decea.mil.br/publicacao/{regulamento}")
    codigo_http = r.status_code
    # print(codigo_http)
    if codigo_http == 200:
        pagina = BeautifulSoup(r.text, 'html.parser')
        informarcao = pagina.find_all(attrs={"class": "d-flex justify-content-between align-items-center"})
        informarcao = trata_data_site(informarcao[0].find_next("strong").text)
        # print(informarcao)
        if informarcao == regulamento_[2]:
            pass
            # print(f"{regulamento_[0]} {regulamento_[1]} está atualizado")
            # print(f"A data de expedição é {informarcao}")
        else:
            # print(f"{regulamento_[0]} {regulamento_[1]} não está atualizado")
            # print(f"A data de expedição é {informarcao} e a versão que tens é de {regulamento_[2]}")
            atualiza_planilha(arq, lista_indices, )
            documento_desatualizado = f"{regulamento_[0]} {regulamento_[1]}"
            # print("Não está atualizado!!!!")
    else:
        print("Regulamento não foi encontrado no sitema AISWEB!!!")
    return documento_desatualizado


def atualiza_planilha(arquivo_, lista_indices):
    # print(lista_indices)
    arquivo_.worksheets[0].cell(row=lista_indices[0], column=lista_indices[1], value="Não")
    # arquivo_.save(r"D:\projetos\pessoal\sistema_paim\Documentação Atualizador Publicações\publicacoes_torre_sdco.xlsx")
    arquivo_.save(r'D:\Jocimar\TWR_SDCO_PROJETOS\publicacoes_torre_sdco.xlsx')


if __name__ == '__main__':
    db_excel = abrir_excel()
    documentos = verificador_documentacao_atualizado(db_excel)
    if len(documentos) > 0:
        print("As seguintes publicações estão desatualizadas!")
        for i in range(len(documentos)):
            print(documentos[i])
    print('Jocimar Braga')
